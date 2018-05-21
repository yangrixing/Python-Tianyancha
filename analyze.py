#!/usr/bin/python
# -*- coding: utf-8 -*-
# @Time    : 2018/2/5 15:56
# @Author  :
# @Site    : 
# @File    : analyze.py

import xlrd
from openpyxl.workbook import Workbook
import arrow


def exceldata(filename, n):
    """
    open excel file read data
    :param filename:  excel file name
    :param n: sheet number
    :return:  xlrd sheet data
    """
    try:
        book = xlrd.open_workbook(filename)
        sheets = book.sheets()
        sheet = sheets[n]
        return sheet
    except Exception as e:
        print(e)
        return None


def cleandata(sheet, datalist, dictname, splitstate):
    """
    clean xlsx data and create dict
    :param sheet: excel sheet
    :param datalist: data row
    :param dictname: dict txt file name
    :param splitstate: rows string splite switch
    :return: None
    """
    data = []
    data_uniq = []
    with open(dictname, "w", encoding="utf-8") as F:
        if splitstate:
            print("处理带分割符信息")
            for datarow in range(sheet.nrows):
                # 处理表头
                if datarow != 0:
                    col = sheet.cell(datarow, datalist).value.split("|,|")
                    for word in col:
                        if word != "" and word != "暂无":
                            data.append(word)
            print("信息去重")
            data_uniq = sorted(set(data), key=data.index)
            for word in data_uniq:
                F.writelines(word + "\n")
        else:
            print("处理不带分隔符信息")
            for datarow in range(sheet.nrows):
                if datarow != 0:
                    col = sheet.cell(datarow, datalist).value
                    if col != "" and col != "暂无":
                        data.append(col)
            print("信息去重")
            data_uniq = sorted(set(data), key=data.index)
            for word in data_uniq:
                F.writelines(word + "\n")
    print("处理完成，关闭文件")
    F.close()


def analyze(dictfile, sheet, row, outputfile, sheetkeys, headers):
    """
    分析关系
    :param dictfile: 字典文件
    :param sheet: Excel数据表
    :param row: 数据行
    :param outputfile: 输出文件名
    :param sheetkeys: List
    :param headers: 标题行
    :return:
    """
    print("开始分析，请稍后……")
    keys = []
    with open(dictfile, "r", encoding="utf-8") as R:
        for dictrow in R.readlines():
            keys.append(dictrow.split("\n")[0])
    R.close()
    wb = Workbook()
    ws = wb.active
    ws.append(headers)

    for dictkey in keys:
        count = 0
        resultlist = []
        resultstr = ""
        for datarow in range(sheet.nrows):
            if datarow != 0:
                col = sheet.cell(datarow, row).value
                if dictkey in col:
                    # print(len(dictkey), len(col))
                    if len(dictkey) == len(col):
                        for rowkey in sheetkeys:
                            resultlist.append(sheet.cell(datarow, rowkey).value) # 获取符合条件的对应列数据
                    else:
                        for word in col.split("|,|"):
                            if dictkey in word:
                                if len(dictkey) == len(word):
                                    for rowkey in sheetkeys:
                                        resultlist.append(sheet.cell(datarow, rowkey).value)  # 获取符合条件的对应列数据
        resultlist = list(set(resultlist))
        count = len(resultlist)
        for resultone in resultlist:
            resultstr += resultone + ","
        result = [dictkey, count, resultstr]
        ws.append(result)
    print("处理完毕，正在生成结果文件……")
    wb.save(filename=outputfile)
    print("处理完成。")


if __name__ == "__main__":
    sheet = exceldata("result.xlsx", 0)
    # 股东信息字典生成
    cleandata(sheet, 16, "share_dict.txt", True)
    # 法人名称字典
    cleandata(sheet, 2, "legal_dict.txt", False)
    # 对外投资字典
    cleandata(sheet, 17, "invest_dict.txt", True)
    # 公司名称
    cleandata(sheet, 0, "cmp_dict.txt", False)
    # 分析
    sheetkeyslist = [0]  # 数据表内
    sheetheader = ['法人名称', '拥有公司数量', '参股公司名单']
    # 参数（字典，Excel表对象，与字典对应的Excel数据列，输出文件名，数据表内需要分析关联的数据列名，输出文件标题行）
    analyze("legal_dict.txt", sheet, 2, "test1.xlsx", sheetkeyslist, sheetheader)
